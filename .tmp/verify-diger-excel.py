# -*- coding: utf-8 -*-
"""Verify the generated Diger excel."""
from openpyxl import load_workbook

wb = load_workbook("Sayt üçün 24082026_260812_162518 -Diger.xlsx")
print("Sheets:", wb.sheetnames)

ws = wb["Məhsullar"]
rows = list(ws.iter_rows(min_row=3, values_only=True))
headers = rows[0]
print("Headers:", headers)
data = rows[1:]
print("Product rows:", len(data))

# sanity checks
counts = {}
cats = {}
conds = {}
prev_ok = True
for r in data:
    if not r[0]:
        continue
    num, model, barcode, cond, qty, price, old, main, sub, brand = r[:10]
    counts.setdefault((main, sub), []).append(model)
    conds[cond] = conds.get(cond, 0) + 1
    if old != round(price + 50, 2):
        prev_ok = False
        print("OLD PRICE MISMATCH", num, price, old)

print("\n-- Category counts --")
for k in sorted(counts):
    print(f"{k[0]} / {k[1]}: {len(counts[k])}")

print("\n-- Condition counts --", conds)
print("Old price = price+50 for all:", prev_ok)

# check specs have all 3 languages and condition line
specs_ok = True
for r in data:
    if not r[0]:
        continue
    num, az, ru, en = r[0], r[10], r[11], r[12]
    if not (az and ru and en):
        specs_ok = False
        print("MISSING SPECS", num)
    for lang, text in (("AZ", az), ("RU", ru), ("EN", en)):
        if "Vəziyyəti" not in (text or "") and "Состояние" not in (text or "") and "Condition" not in (text or ""):
            specs_ok = False
            print("MISSING CONDITION LINE", num, lang)
print("All specs present with condition:", specs_ok)

ws2 = wb["Kateqoriyalar"]
print("\n-- Categories sheet --")
for r in ws2.iter_rows(min_row=3, max_col=5, values_only=True):
    if r[0]:
        print(r)
