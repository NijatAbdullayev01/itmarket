# -*- coding: utf-8 -*-
"""Generate the Diger məhsullar Excel (31.08.2026) for the IT Market site.

Source: "Sayt üçün 31082026- DIGER_260831_110926.docx" + live catalog categories.
- Old price = discounted price + 55 AZN.
- Condition: Yeni (products marked "yeni") / İşlənmiş (all others).
- RU/EN feature columns are for the translation dictionary database.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAIN_NETWORK = "Şəbəkə avadanlıqları"
MAIN_SERVER = "Server"
MAIN_UPS = "UPS"

COND_YENI = "Yeni"
COND_ISLENMIS = "İşlənmiş"

C_NEW = ("Vəziyyəti: Yeni (istifadə olunmamış, orijinal qablaşdırmada)",
         "Состояние: Новый (не использовался, в оригинальной упаковке)",
         "Condition: New (unused, original packaging)")
C_USED = ("Vəziyyəti: İşlənmiş (tam saz, texniki test edilmiş)",
          "Состояние: Б/У (в полном рабочем состоянии, протестировано)",
          "Condition: Used (fully functional, tested)")


def condition_specs(condition):
    return C_NEW if condition == COND_YENI else C_USED


def combine(cond, az, ru, en):
    c_az, c_ru, c_en = condition_specs(cond)
    return [c_az, *az], [c_ru, *ru], [c_en, *en]


P = []
IMAGES = {}


def add(num, model, barcode, qty, price, main, sub, brand, cond, az, ru, en, img=""):
    P.append((num, model, barcode, qty, price, main, sub, brand, cond,
              *combine(cond, az, ru, en)))
    if img:
        IMAGES[num] = img


# Sub-categories
SUB_SW = "Kommutator"
SUB_ROUTER = "Router"
SUB_VOIP = "VoIP və IP telefonlar"
SUB_NET_ACC = "Şəbəkə aksesuarları"
SUB_SFP = "SFP modullar"
SUB_NET_PSU = "Şəbəkə enerji təchizatı"
SUB_SRV_ACC = "Server aksesuarları"
SUB_SRV_PSU = "Server enerji təchizatı"
SUB_HDD = "HDD"
SUB_SRV_NIC = "Şəbəkə adapteri"
SUB_UPS_ACC = "UPS aksesuarları"

# ================================================================== KOMMUTATOR

add(1, "HP 2510 48 J9020A", None, 2, 200, MAIN_NETWORK, SUB_SW, "HP", COND_ISLENMIS,
    ["Tip: İdarə olunan L2 kommutator (HP ProCurve 2510-48)",
     "Portlar: 48× 10/100 Mbps + 2× 10/100/1000 + 2× SFP (mini-GBIC)",
     "Kommutasiya tutumu: 17.6 Gbps",
     "Paket ötürmə: 13 Mpps",
     "VLAN / QoS / Link Aggregation dəstəyi",
     "Menecment: CLI / Web / SNMP",
     "Korpus: 1U, rack-mount"],
    ["Тип: Управляемый L2 коммутатор (HP ProCurve 2510-48)",
     "Порты: 48× 10/100 Мбит/с + 2× 10/100/1000 + 2× SFP (mini-GBIC)",
     "Коммутационная способность: 17.6 Гбит/с",
     "Скорость пакетов: 13 млн пакетов/с",
     "Поддержка VLAN / QoS / Link Aggregation",
     "Управление: CLI / Web / SNMP",
     "Корпус: 1U, монтаж в стойку"],
    ["Type: Managed L2 switch (HP ProCurve 2510-48)",
     "Ports: 48× 10/100 Mbps + 2× 10/100/1000 + 2× SFP (mini-GBIC)",
     "Switching capacity: 17.6 Gbps",
     "Forwarding rate: 13 Mpps",
     "VLAN / QoS / Link Aggregation support",
     "Management: CLI / Web / SNMP",
     "Chassis: 1U, rack-mountable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/95092/704720/J9020AABA__04097.1784347799.jpg?c=1")

add(2, "HP 2610-24/12PWR J9086A", None, 4, 230, MAIN_NETWORK, SUB_SW, "HP", COND_ISLENMIS,
    ["Tip: İdarə olunan L2 kommutator (HP ProCurve 2610-24/12PWR)",
     "Portlar: 24× 10/100 Mbps (12 portda PoE) + 2× 10/100/1000 + 2× SFP",
     "PoE: IEEE 802.3af (12 porta qədər 15.4 W)",
     "Menecment: CLI / Web / SNMP",
     "VLAN / QoS / ACL dəstəyi",
     "Korpus: 1U, rack-mount"],
    ["Тип: Управляемый L2 коммутатор (HP ProCurve 2610-24/12PWR)",
     "Порты: 24× 10/100 Мбит/с (PoE на 12 портах) + 2× 10/100/1000 + 2× SFP",
     "PoE: IEEE 802.3af (до 15.4 Вт на 12 портов)",
     "Управление: CLI / Web / SNMP",
     "Поддержка VLAN / QoS / ACL",
     "Корпус: 1U, монтаж в стойку"],
    ["Type: Managed L2 switch (HP ProCurve 2610-24/12PWR)",
     "Ports: 24× 10/100 Mbps (PoE on 12 ports) + 2× 10/100/1000 + 2× SFP",
     "PoE: IEEE 802.3af (up to 15.4 W on 12 ports)",
     "Management: CLI / Web / SNMP",
     "VLAN / QoS / ACL support",
     "Chassis: 1U, rack-mountable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/270223/706019/J9086A-o__94655.1784580281.jpg?c=1")

add(3, "AMX NXA-ENET8-2POE", None, 3, 240, MAIN_NETWORK, SUB_SW, "AMX", COND_ISLENMIS,
    ["Tip: PoE kommutator (AMX Enova NXA-ENET8-2POE)",
     "Portlar: 8× 10/100/1000 Mbps (2 portda PoE)",
     "AV şəbəkələri və idarəetmə sistemləri üçün",
     "PoE: IEEE 802.3af",
     "Korpus: Desktop / rack-mount"],
    ["Тип: PoE-коммутатор (AMX Enova NXA-ENET8-2POE)",
     "Порты: 8× 10/100/1000 Мбит/с (PoE на 2 портах)",
     "Для AV-сетей и систем управления",
     "PoE: IEEE 802.3af",
     "Корпус: Настольный / монтаж в стойку"],
    ["Type: PoE switch (AMX Enova NXA-ENET8-2POE)",
     "Ports: 8× 10/100/1000 Mbps (PoE on 2 ports)",
     "For AV networks and control systems",
     "PoE: IEEE 802.3af",
     "Chassis: Desktop / rack-mountable"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/352453/351921/108288631356857.1754678771.jpg?c=2")

add(4, "NETGEAR FS105", None, 32, 9, MAIN_NETWORK, SUB_SW, "NETGEAR", COND_ISLENMIS,
    ["Tip: İdarəolunmaz Fast Ethernet kommutator (NETGEAR FS105)",
     "Portlar: 5× 10/100 Mbps",
     "Plug-and-play, Auto MDI/MDIX",
     "Kompakt desktop korpus"],
    ["Тип: Неуправляемый Fast Ethernet коммутатор (NETGEAR FS105)",
     "Порты: 5× 10/100 Мбит/с",
     "Plug-and-play, Auto MDI/MDIX",
     "Компактный настольный корпус"],
    ["Type: Unmanaged Fast Ethernet switch (NETGEAR FS105)",
     "Ports: 5× 10/100 Mbps",
     "Plug-and-play, Auto MDI/MDIX",
     "Compact desktop chassis"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/352453/351921/108288631356857.1754678771.jpg?c=2")

add(5, "NETGEAR GS108E", None, 11, 35, MAIN_NETWORK, SUB_SW, "NETGEAR", COND_ISLENMIS,
    ["Tip: Smart (idarə olunan) Gigabit kommutator (NETGEAR ProSAFE GS108E)",
     "Portlar: 8× 10/100/1000 Mbps",
     "VLAN, QoS, IGMP Snooping, Loop Detection",
     "Green Ethernet (enerjiyə qənaət)",
     "Fanless metal korpus",
     "Menecment: Web (ProSAFE Plus Utility)"],
    ["Тип: Smart управляемый гигабитный коммутатор (NETGEAR ProSAFE GS108E)",
     "Порты: 8× 10/100/1000 Мбит/с",
     "VLAN, QoS, IGMP Snooping, Loop Detection",
     "Green Ethernet (энергосбережение)",
     "Металлический корпус без вентилятора",
     "Управление: Web (ProSAFE Plus Utility)"],
    ["Type: Smart managed Gigabit switch (NETGEAR ProSAFE GS108E)",
     "Ports: 8× 10/100/1000 Mbps",
     "VLAN, QoS, IGMP Snooping, Loop Detection",
     "Green Ethernet (energy saving)",
     "Fanless metal chassis",
     "Management: Web (ProSAFE Plus Utility)"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/352453/351921/108288631356857.1754678771.jpg?c=2")

add(6, "NETGEAR FS108P", None, 7, 30, MAIN_NETWORK, SUB_SW, "NETGEAR", COND_ISLENMIS,
    ["Tip: İdarəolunmaz Fast Ethernet PoE kommutator (NETGEAR FS108P)",
     "Portlar: 8× 10/100 Mbps (4 portda PoE)",
     "PoE: IEEE 802.3af",
     "Kompakt desktop korpus"],
    ["Тип: Неуправляемый Fast Ethernet PoE коммутатор (NETGEAR FS108P)",
     "Порты: 8× 10/100 Мбит/с (PoE на 4 портах)",
     "PoE: IEEE 802.3af",
     "Компактный настольный корпус"],
    ["Type: Unmanaged Fast Ethernet PoE switch (NETGEAR FS108P)",
     "Ports: 8× 10/100 Mbps (PoE on 4 ports)",
     "PoE: IEEE 802.3af",
     "Compact desktop chassis"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/352453/351921/108288631356857.1754678771.jpg?c=2")

add(7, "NETGEAR FS728TP", None, 1, 270, MAIN_NETWORK, SUB_SW, "NETGEAR", COND_ISLENMIS,
    ["Tip: Smart PoE kommutator (NETGEAR ProSAFE FS728TP)",
     "Portlar: 24× 10/100 Mbps (PoE) + 2× 10/100/1000 + 2× combo SFP",
     "PoE büdcəsi: 192 W (IEEE 802.3af)",
     "Kommutasiya tutumu: 12.8 Gbps",
     "VLAN / QoS / ACL / IGMP Snooping",
     "Menecment: Web"],
    ["Тип: Smart PoE-коммутатор (NETGEAR ProSAFE FS728TP)",
     "Порты: 24× 10/100 Мбит/с (PoE) + 2× 10/100/1000 + 2× combo SFP",
     "PoE-бюджет: 192 Вт (IEEE 802.3af)",
     "Коммутационная способность: 12.8 Гбит/с",
     "VLAN / QoS / ACL / IGMP Snooping",
     "Управление: Web"],
    ["Type: Smart PoE switch (NETGEAR ProSAFE FS728TP)",
     "Ports: 24× 10/100 Mbps (PoE) + 2× 10/100/1000 + 2× combo SFP",
     "PoE budget: 192 W (IEEE 802.3af)",
     "Switching capacity: 12.8 Gbps",
     "VLAN / QoS / ACL / IGMP Snooping",
     "Management: Web"],
    img="https://cdn11.bigcommerce.com/s-2mzfab8ce4/images/stencil/1280x1280/products/261801/383680/yhst-172536426-9_2519_598709620__50965.1527506156__53981.1775500293.jpg?c=1")

add(8, "NETGEAR GS110TP", None, 1, 40, MAIN_NETWORK, SUB_SW, "NETGEAR", COND_ISLENMIS,
    ["Tip: Smart Gigabit PoE kommutator (NETGEAR ProSAFE GS110TP)",
     "Portlar: 8× 10/100/1000 Mbps (PoE) + 2× SFP",
     "PoE: IEEE 802.3af",
     "VLAN / QoS / IGMP Snooping",
     "Menecment: Web"],
    ["Тип: Smart гигабитный PoE-коммутатор (NETGEAR ProSAFE GS110TP)",
     "Порты: 8× 10/100/1000 Мбит/с (PoE) + 2× SFP",
     "PoE: IEEE 802.3af",
     "VLAN / QoS / IGMP Snooping",
     "Управление: Web"],
    ["Type: Smart Gigabit PoE switch (NETGEAR ProSAFE GS110TP)",
     "Ports: 8× 10/100/1000 Mbps (PoE) + 2× SFP",
     "PoE: IEEE 802.3af",
     "VLAN / QoS / IGMP Snooping",
     "Management: Web"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/352453/351921/108288631356857.1754678771.jpg?c=2")

add(19, "ARUBA S3500-24-US 4×10G", None, 1, 450, MAIN_NETWORK, SUB_SW, "Aruba", COND_ISLENMIS,
    ["Tip: Mobility Access Switch (Aruba S3500-24T, L2/L3)",
     "Portlar: 24× 10/100/1000 Mbps + uplink modul slotu",
     "Uplink: 4× 10G SFP+ (modul ilə)",
     "Kommutasiya tutumu: 128 Gbps",
     "Paket ötürmə: 95 Mpps",
     "Menecment: Web / CLI"],
    ["Тип: Mobility Access Switch (Aruba S3500-24T, L2/L3)",
     "Порты: 24× 10/100/1000 Мбит/с + слот аплинк-модуля",
     "Uplink: 4× 10G SFP+ (с модулем)",
     "Коммутационная способность: 128 Гбит/с",
     "Скорость пакетов: 95 млн пакетов/с",
     "Управление: Web / CLI"],
    ["Type: Mobility Access Switch (Aruba S3500-24T, L2/L3)",
     "Ports: 24× 10/100/1000 Mbps + uplink module slot",
     "Uplink: 4× 10G SFP+ (with module)",
     "Switching capacity: 128 Gbps",
     "Forwarding rate: 95 Mpps",
     "Management: Web / CLI"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/284163/717562/24-Network-Switches-Gen__80346.1784339653.jpg?c=1")

add(20, "ARUBA S2500 48P 4×10G", None, 4, 550, MAIN_NETWORK, SUB_SW, "Aruba", COND_ISLENMIS,
    ["Tip: Mobility Access Switch (Aruba S2500-48P, L2+)",
     "Portlar: 48× 10/100/1000 Mbps PoE+ + 4× 10G SFP+",
     "PoE büdcəsi: 580 W (port başına 30 W-a qədər)",
     "Menecment: Web / CLI",
     "Korpus: 1U, rack-mount"],
    ["Тип: Mobility Access Switch (Aruba S2500-48P, L2+)",
     "Порты: 48× 10/100/1000 Мбит/с PoE+ + 4× 10G SFP+",
     "PoE-бюджет: 580 Вт (до 30 Вт на порт)",
     "Управление: Web / CLI",
     "Корпус: 1U, монтаж в стойку"],
    ["Type: Mobility Access Switch (Aruba S2500-48P, L2+)",
     "Ports: 48× 10/100/1000 Mbps PoE+ + 4× 10G SFP+",
     "PoE budget: 580 W (up to 30 W per port)",
     "Management: Web / CLI",
     "Chassis: 1U, rack-mountable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/520259/708156/JW670A__04319.1784339643.jpg?c=1")

add(21, "ARUBA S3500-24F 4×10G", None, 1, 600, MAIN_NETWORK, SUB_SW, "Aruba", COND_ISLENMIS,
    ["Tip: Mobility Access Switch (Aruba S3500-24F, SFP fiber)",
     "Portlar: 24× 1000BASE-X SFP + 4× 10G SFP+",
     "Kommutasiya tutumu: 128 Gbps",
     "Menecment: Web / CLI",
     "Korpus: 1U, rack-mount"],
    ["Тип: Mobility Access Switch (Aruba S3500-24F, SFP оптика)",
     "Порты: 24× 1000BASE-X SFP + 4× 10G SFP+",
     "Коммутационная способность: 128 Гбит/с",
     "Управление: Web / CLI",
     "Корпус: 1U, монтаж в стойку"],
    ["Type: Mobility Access Switch (Aruba S3500-24F, SFP fiber)",
     "Ports: 24× 1000BASE-X SFP + 4× 10G SFP+",
     "Switching capacity: 128 Gbps",
     "Management: Web / CLI",
     "Chassis: 1U, rack-mountable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/284163/717562/24-Network-Switches-Gen__80346.1784339653.jpg?c=1")

add(22, "HP 5500 SERIES SWITCH JG542A", None, 7, 600, MAIN_NETWORK, SUB_SW, "HP", COND_ISLENMIS,
    ["Tip: L3 kommutator (HP 5500-48G-PoE+-4SFP HI, JG542A)",
     "Portlar: 48× 10/100/1000 Mbps PoE+ + 4× SFP + 2× SFP+",
     "Kommutasiya tutumu: 224 Gbps",
     "Paket ötürmə: 166.6 Mpps",
     "PoE büdcəsi: 1440 W",
     "Statik/dinamik rutinq (OSPF, BGP, RIP)",
     "Menecment: CLI / Web / SNMP",
     "Korpus: 1U, rack-mount"],
    ["Тип: L3 коммутатор (HP 5500-48G-PoE+-4SFP HI, JG542A)",
     "Порты: 48× 10/100/1000 Мбит/с PoE+ + 4× SFP + 2× SFP+",
     "Коммутационная способность: 224 Гбит/с",
     "Скорость пакетов: 166.6 млн пакетов/с",
     "PoE-бюджет: 1440 Вт",
     "Статическая/динамическая маршрутизация (OSPF, BGP, RIP)",
     "Управление: CLI / Web / SNMP",
     "Корпус: 1U, монтаж в стойку"],
    ["Type: L3 switch (HP 5500-48G-PoE+-4SFP HI, JG542A)",
     "Ports: 48× 10/100/1000 Mbps PoE+ + 4× SFP + 2× SFP+",
     "Switching capacity: 224 Gbps",
     "Forwarding rate: 166.6 Mpps",
     "PoE budget: 1440 W",
     "Static/dynamic routing (OSPF, BGP, RIP)",
     "Management: CLI / Web / SNMP",
     "Chassis: 1U, rack-mountable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/288426/708546/JG542A-o__04973.1784331910.jpg?c=1")

add(23, "H3C S3600 JD327A SWITCH", None, 3, 230, MAIN_NETWORK, SUB_SW, "H3C", COND_ISLENMIS,
    ["Tip: İdarə olunan L2/L3 kommutator (H3C S3600-28P-EI)",
     "Portlar: 24× 10/100 Mbps + 4× 1000Base-X SFP + Console",
     "Kommutasiya tutumu: 32 Gbps",
     "Paket ötürmə: 9.6 Mpps",
     "MAC cədvəli: 16K",
     "Menecment: CLI / Web / SNMP",
     "Korpus: 1U, rack-mount"],
    ["Тип: Управляемый L2/L3 коммутатор (H3C S3600-28P-EI)",
     "Порты: 24× 10/100 Мбит/с + 4× 1000Base-X SFP + Console",
     "Коммутационная способность: 32 Гбит/с",
     "Скорость пакетов: 9.6 млн пакетов/с",
     "Таблица MAC: 16K",
     "Управление: CLI / Web / SNMP",
     "Корпус: 1U, монтаж в стойку"],
    ["Type: Managed L2/L3 switch (H3C S3600-28P-EI)",
     "Ports: 24× 10/100 Mbps + 4× 1000Base-X SFP + Console",
     "Switching capacity: 32 Gbps",
     "Forwarding rate: 9.6 Mpps",
     "MAC table: 16K",
     "Management: CLI / Web / SNMP",
     "Chassis: 1U, rack-mountable"])

add(74, "CISCO DS-C9124-K9 SFP SWITCH", None, 4, 550, MAIN_NETWORK, SUB_SW, "Cisco", COND_ISLENMIS,
    ["Tip: Fibre Channel (SAN) kommutator (Cisco MDS 9124)",
     "Portlar: 8× 4 Gbps aktiv + 16× SFP (Ports on Demand)",
     "Sürətlər: 1/2/4 Gbps FC",
     "Supervisor-2 modulu",
     "Menecment: NX-OS / CLI",
     "Korpus: 1U, rack-mount"],
    ["Тип: Коммутатор Fibre Channel (SAN) (Cisco MDS 9124)",
     "Порты: 8× 4 Гбит/с активных + 16× SFP (Ports on Demand)",
     "Скорости: 1/2/4 Гбит/с FC",
     "Модуль Supervisor-2",
     "Управление: NX-OS / CLI",
     "Корпус: 1U, монтаж в стойку"],
    ["Type: Fibre Channel (SAN) switch (Cisco MDS 9124)",
     "Ports: 8× 4 Gbps active + 16× SFP (Ports on Demand)",
     "Speeds: 1/2/4 Gbps FC",
     "Supervisor-2 module",
     "Management: NX-OS / CLI",
     "Chassis: 1U, rack-mountable"],
    img="https://cdn11.bigcommerce.com/s-z8mygza4ws/images/stencil/1280x1280/products/91082/1091722/DS-C9124-1-K9__24646.1712647447.jpg?c=1")

# ============================================================ VoIP və IP telefonlar

add(11, "CISCO IP PHONE 7916 EXPANSION MODULE", None, 1, 75, MAIN_NETWORK, SUB_VOIP, "Cisco", COND_ISLENMIS,
    ["Tip: IP telefon genişləndirmə modulu (Cisco 7916)",
     "Düymələr: 12 fiziki + 12 səhifə (cəmi 24)",
     "Ekran: Böyük LCD",
     "Uyğunluq: Cisco 7962G, 7965G, 7975G",
     "Maks. 2 modul bir telefonda"],
    ["Тип: Модуль расширения IP-телефона (Cisco 7916)",
     "Кнопки: 12 физических + 12 по странице (всего 24)",
     "Экран: Большой LCD",
     "Совместимость: Cisco 7962G, 7965G, 7975G",
     "До 2 модулей на телефон"],
    ["Type: IP phone expansion module (Cisco 7916)",
     "Buttons: 12 physical + 12 page (24 total)",
     "Display: Large LCD",
     "Compatibility: Cisco 7962G, 7965G, 7975G",
     "Up to 2 modules per phone"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/353386/352974/1060969402__85281.1755217335.jpg?c=2")

add(13, "CISCO CP8800-A-KEM EXPANSION MODULE YENI", None, 1, 300, MAIN_NETWORK, SUB_VOIP, "Cisco", COND_YENI,
    ["Tip: IP telefon genişləndirmə modulu (Cisco 8800 Key Expansion Module)",
     "Düymələr: 14 fiziki + 14 səhifə (cəmi 28)",
     "Ekran: 3.5\" rəngli TFT (480×272)",
     "Uyğunluq: Cisco 8851, 8861, 8865",
     "Maks. 3 modul (8861/8865)"],
    ["Тип: Модуль расширения IP-телефона (Cisco 8800 Key Expansion Module)",
     "Кнопки: 14 физических + 14 по странице (всего 28)",
     "Экран: 3.5\" цветной TFT (480×272)",
     "Совместимость: Cisco 8851, 8861, 8865",
     "До 3 модулей (8861/8865)"],
    ["Type: IP phone key expansion module (Cisco 8800 KEM)",
     "Buttons: 14 physical + 14 page (28 total)",
     "Display: 3.5\" color TFT (480×272)",
     "Compatibility: Cisco 8851, 8861, 8865",
     "Up to 3 modules (8861/8865)"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/353386/352974/1060969402__85281.1755217335.jpg?c=2")

add(14, "CISCO CP-SINGLFOOTSTAND FOR 7914,15 OR 16  YENI", None, 1, 90, MAIN_NETWORK, SUB_VOIP, "Cisco", COND_YENI,
    ["Tip: Telefon ayaq dayağı (tək modul üçün)",
     "Uyğunluq: Cisco 7914, 7915, 7916 genişləndirmə modulları",
     "Material: Plastik"],
    ["Тип: Подставка для телефона (для одного модуля)",
     "Совместимость: модули расширения Cisco 7914, 7915, 7916",
     "Материал: Пластик"],
    ["Type: Phone footstand (single module)",
     "Compatibility: Cisco 7914, 7915, 7916 expansion modules",
     "Material: Plastic"])

add(16, "CISCO CPDSKCH-8821-BUN DESKTOP CHARGER YENI", None, 1, 540, MAIN_NETWORK, SUB_VOIP, "Cisco", COND_YENI,
    ["Tip: Stolüstü şarj cihazı (Cisco 8821 üçün)",
     "Uyğunluq: Cisco 8821 simsiz IP telefon",
     "Batareyanı şarj edir"],
    ["Тип: Настольное зарядное устройство (для Cisco 8821)",
     "Совместимость: беспроводной IP-телефон Cisco 8821",
     "Зарядка аккумулятора"],
    ["Type: Desktop charger (for Cisco 8821)",
     "Compatibility: Cisco 8821 wireless IP phone",
     "Charges the battery"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/19438/18217/1035446513__70135.1772573885.jpg?c=2")

add(17, "CISCO CP3905-HS HEADSET YENI", None, 10, 50, MAIN_NETWORK, SUB_VOIP, "Cisco", COND_YENI,
    ["Tip: IP telefon qulaqlığı (Cisco 3905 üçün)",
     "Qoşulma: RJ-9",
     "Səs: Genişzolaqlı audio"],
    ["Тип: Гарнитура для IP-телефона (для Cisco 3905)",
     "Подключение: RJ-9",
     "Аудио: Широкополосный звук"],
    ["Type: IP phone headset (for Cisco 3905)",
     "Connection: RJ-9",
     "Audio: Wideband audio"],
    img="https://cdn11.bigcommerce.com/s-2mzfab8ce4/images/stencil/1280x1280/products/383393/367089/_7TVQ4hNOUaRMzj7qneeEQ.c-r__41615__82535__53686.1723820935.jpg?c=1")

add(75, "CISCO HEADSET CABLE", None, 60, 9, MAIN_NETWORK, SUB_VOIP, "Cisco", COND_ISLENMIS,
    ["Tip: Qulaqlıq kabeli (Cisco IP telefonlar üçün)",
     "Qoşulma: RJ-9 / aux"],
    ["Тип: Кабель гарнитуры (для IP-телефонов Cisco)",
     "Подключение: RJ-9 / aux"],
    ["Type: Headset cable (for Cisco IP phones)",
     "Connection: RJ-9 / aux"],
    img="https://cdn11.bigcommerce.com/s-2mzfab8ce4/images/stencil/1280x1280/products/383393/367089/_7TVQ4hNOUaRMzj7qneeEQ.c-r__41615__82535__53686.1723820935.jpg?c=1")

# ===================================================================== ROUTER (modullar)

add(15, "CISCO EHWIC-4ESG-P YENI", None, 1, 110, MAIN_NETWORK, SUB_ROUTER, "Cisco", COND_YENI,
    ["Tip: EtherSwitch xidmət modulu (Cisco EHWIC-4ESG-P)",
     "Portlar: 4× 10/100/1000 Mbps (PoE)",
     "Uyğunluq: Cisco ISR G2 routerlər (1900/2900/3900)",
     "PoE: IEEE 802.3af"],
    ["Тип: Сервисный модуль EtherSwitch (Cisco EHWIC-4ESG-P)",
     "Порты: 4× 10/100/1000 Мбит/с (PoE)",
     "Совместимость: маршрутизаторы Cisco ISR G2 (1900/2900/3900)",
     "PoE: IEEE 802.3af"],
    ["Type: EtherSwitch service module (Cisco EHWIC-4ESG-P)",
     "Ports: 4× 10/100/1000 Mbps (PoE)",
     "Compatibility: Cisco ISR G2 routers (1900/2900/3900)",
     "PoE: IEEE 802.3af"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(33, "CISCO VWIC3-2MFT-T1/E1 VOICE WAN CARD", None, 2, 78, MAIN_NETWORK, SUB_ROUTER, "Cisco", COND_ISLENMIS,
    ["Tip: Voice WAN interfeys kartı (Cisco VWIC3-2MFT-T1/E1)",
     "Portlar: 2× T1/E1",
     "Uyğunluq: Cisco ISR G2 routerlər",
     "Səs/Data dəstəyi"],
    ["Тип: Голосовая WAN-карта (Cisco VWIC3-2MFT-T1/E1)",
     "Порты: 2× T1/E1",
     "Совместимость: маршрутизаторы Cisco ISR G2",
     "Поддержка голоса/данных"],
    ["Type: Voice WAN interface card (Cisco VWIC3-2MFT-T1/E1)",
     "Ports: 2× T1/E1",
     "Compatibility: Cisco ISR G2 routers",
     "Voice/data support"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(34, "CISCO EHWIC-4ESG MODULE ISR", None, 3, 85, MAIN_NETWORK, SUB_ROUTER, "Cisco", COND_ISLENMIS,
    ["Tip: EtherSwitch xidmət modulu (Cisco EHWIC-4ESG)",
     "Portlar: 4× 10/100/1000 Mbps",
     "Uyğunluq: Cisco ISR G2 routerlər"],
    ["Тип: Сервисный модуль EtherSwitch (Cisco EHWIC-4ESG)",
     "Порты: 4× 10/100/1000 Мбит/с",
     "Совместимость: маршрутизаторы Cisco ISR G2"],
    ["Type: EtherSwitch service module (Cisco EHWIC-4ESG)",
     "Ports: 4× 10/100/1000 Mbps",
     "Compatibility: Cisco ISR G2 routers"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(36, "CISCO EHWIC-4G-LTE-VZ CARD", None, 2, 80, MAIN_NETWORK, SUB_ROUTER, "Cisco", COND_ISLENMIS,
    ["Tip: 4G LTE EHWIC modulu (Cisco EHWIC-4G-LTE-VZ)",
     "Əlaqə: 4G LTE (Verizon)",
     "Uyğunluq: Cisco ISR G2 routerlər"],
    ["Тип: Модуль 4G LTE EHWIC (Cisco EHWIC-4G-LTE-VZ)",
     "Связь: 4G LTE (Verizon)",
     "Совместимость: маршрутизаторы Cisco ISR G2"],
    ["Type: 4G LTE EHWIC module (Cisco EHWIC-4G-LTE-VZ)",
     "Connectivity: 4G LTE (Verizon)",
     "Compatibility: Cisco ISR G2 routers"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(67, "CISCO ROUTER RAM PVDM3-16 V01", None, 2, 65, MAIN_NETWORK, SUB_ROUTER, "Cisco", COND_ISLENMIS,
    ["Tip: DSP (səs) modulu (Cisco PVDM3-16)",
     "Səs kanalları: 16",
     "Uyğunluq: Cisco ISR G2 routerlər",
     "Kodaklar: G.711, G.729, G.722"],
    ["Тип: DSP-модуль (голос) (Cisco PVDM3-16)",
     "Голосовых каналов: 16",
     "Совместимость: маршрутизаторы Cisco ISR G2",
     "Кодеки: G.711, G.729, G.722"],
    ["Type: DSP (voice) module (Cisco PVDM3-16)",
     "Voice channels: 16",
     "Compatibility: Cisco ISR G2 routers",
     "Codecs: G.711, G.729, G.722"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

# ========================================================= ŞƏBƏKƏ AKSESUARLARI (modullar, stack, adapterlər)

add(24, "CISCO C3850-NM-2-10G", None, 24, 200, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Şəbəkə modulu (Cisco Catalyst 3850 üçün)",
     "Portlar: 2× 1G SFP + 2× 10G SFP+",
     "Uyğunluq: Cisco Catalyst 3850 seriyası",
     "Hot-swappable"],
    ["Тип: Сетевой модуль (для Cisco Catalyst 3850)",
     "Порты: 2× 1G SFP + 2× 10G SFP+",
     "Совместимость: серия Cisco Catalyst 3850",
     "Горячая замена"],
    ["Type: Network module (for Cisco Catalyst 3850)",
     "Ports: 2× 1G SFP + 2× 10G SFP+",
     "Compatibility: Cisco Catalyst 3850 series",
     "Hot-swappable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(25, "CISCO C3850-NM-4-10G", None, 3, 220, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Şəbəkə modulu (Cisco Catalyst 3850 üçün)",
     "Portlar: 4× 10G SFP+ (və ya 4× 1G SFP)",
     "Uyğunluq: Cisco Catalyst 3850 seriyası"],
    ["Тип: Сетевой модуль (для Cisco Catalyst 3850)",
     "Порты: 4× 10G SFP+ (или 4× 1G SFP)",
     "Совместимость: серия Cisco Catalyst 3850"],
    ["Type: Network module (for Cisco Catalyst 3850)",
     "Ports: 4× 10G SFP+ (or 4× 1G SFP)",
     "Compatibility: Cisco Catalyst 3850 series"],
    img="https://cdn11.bigcommerce.com/s-ty95hwdn/images/stencil/1280x1280/products/13411/15439/C3850-NM-4-10G__68136.1708635755.jpg?c=2")

add(26, "CISCO C3850-NM-4-1G", None, 5, 150, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Şəbəkə modulu (Cisco Catalyst 3850 üçün)",
     "Portlar: 4× 1G SFP",
     "Uyğunluq: Cisco Catalyst 3850 seriyası"],
    ["Тип: Сетевой модуль (для Cisco Catalyst 3850)",
     "Порты: 4× 1G SFP",
     "Совместимость: серия Cisco Catalyst 3850"],
    ["Type: Network module (for Cisco Catalyst 3850)",
     "Ports: 4× 1G SFP",
     "Compatibility: Cisco Catalyst 3850 series"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(30, "CISCO C2960S STACK V02", None, 5, 30, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Stack modulu (Cisco Catalyst 2960-S üçün)",
     "FlexStack texnologiyası",
     "Uyğunluq: Cisco Catalyst 2960-S seriyası"],
    ["Тип: Модуль стекирования (для Cisco Catalyst 2960-S)",
     "Технология FlexStack",
     "Совместимость: серия Cisco Catalyst 2960-S"],
    ["Type: Stacking module (for Cisco Catalyst 2960-S)",
     "FlexStack technology",
     "Compatibility: Cisco Catalyst 2960-S series"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(31, "CISCO C2960X STACK V02", None, 1, 95, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Stack modulu (Cisco Catalyst 2960-X üçün)",
     "FlexStack-Plus texnologiyası",
     "Uyğunluq: Cisco Catalyst 2960-X seriyası"],
    ["Тип: Модуль стекирования (для Cisco Catalyst 2960-X)",
     "Технология FlexStack-Plus",
     "Совместимость: серия Cisco Catalyst 2960-X"],
    ["Type: Stacking module (for Cisco Catalyst 2960-X)",
     "FlexStack-Plus technology",
     "Compatibility: Cisco Catalyst 2960-X series"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

add(37, "CISCO C3KX-NM-10G", None, 19, 70, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Şəbəkə modulu (Cisco Catalyst 3K-X üçün)",
     "Portlar: 2× SFP+ 10G + 2× SFP 1G",
     "Uyğunluq: Cisco Catalyst 3560-X / 3750-X"],
    ["Тип: Сетевой модуль (для Cisco Catalyst 3K-X)",
     "Порты: 2× SFP+ 10G + 2× SFP 1G",
     "Совместимость: Cisco Catalyst 3560-X / 3750-X"],
    ["Type: Network module (for Cisco Catalyst 3K-X)",
     "Ports: 2× SFP+ 10G + 2× SFP 1G",
     "Compatibility: Cisco Catalyst 3560-X / 3750-X"],
    img="https://img.genuinemodules.com/cache/catalog/products/C3KX-NM-10G/C3KX-NM-10G-1-800x800.jpg")

add(38, "CISCO C3KX-NM- 1G", None, 13, 40, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Şəbəkə modulu (Cisco Catalyst 3K-X üçün)",
     "Portlar: 4× SFP 1G",
     "Uyğunluq: Cisco Catalyst 3560-X / 3750-X"],
    ["Тип: Сетевой модуль (для Cisco Catalyst 3K-X)",
     "Порты: 4× SFP 1G",
     "Совместимость: Cisco Catalyst 3560-X / 3750-X"],
    ["Type: Network module (for Cisco Catalyst 3K-X)",
     "Ports: 4× SFP 1G",
     "Compatibility: Cisco Catalyst 3560-X / 3750-X"],
    img="https://img.genuinemodules.com/cache/catalog/products/C3KX-NM-10G/C3KX-NM-10G-1-800x800.jpg")

add(39, "ARUBA S3500-4×10G", None, 2, 140, MAIN_NETWORK, SUB_NET_ACC, "Aruba", COND_ISLENMIS,
    ["Tip: Uplink modulu (Aruba S3500 üçün)",
     "Portlar: 4× 10G SFP+",
     "Uyğunluq: Aruba S3500 seriyası"],
    ["Тип: Аплинк-модуль (для Aruba S3500)",
     "Порты: 4× 10G SFP+",
     "Совместимость: серия Aruba S3500"],
    ["Type: Uplink module (for Aruba S3500)",
     "Ports: 4× 10G SFP+",
     "Compatibility: Aruba S3500 series"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/410557/968944/cisco_c3850-nm-2-10g_zoom__75756.1784580403.jpg?c=1")

# ============================================================ ŞƏBƏKƏ ADAPTERİ (server)

add(32, "HP NC523SFP 10GBE SFP+ SERVER ADAPTER", None, 4, 120, MAIN_SERVER, SUB_SRV_NIC, "HP", COND_ISLENMIS,
    ["Tip: Şəbəkə adapteri (HP NC523SFP, 10GbE)",
     "Portlar: 2× 10G SFP+",
     "Kontroller: QLogic cLOM8214",
     "İnterfeys: PCIe Gen 2 (x8)",
     "Uyğunluq: HP ProLiant serverlər",
     "Teaming / VLAN / Jumbo Frame dəstəyi"],
    ["Тип: Сетевой адаптер (HP NC523SFP, 10GbE)",
     "Порты: 2× 10G SFP+",
     "Контроллер: QLogic cLOM8214",
     "Интерфейс: PCIe Gen 2 (x8)",
     "Совместимость: серверы HP ProLiant",
     "Поддержка Teaming / VLAN / Jumbo Frame"],
    ["Type: Network adapter (HP NC523SFP, 10GbE)",
     "Ports: 2× 10G SFP+",
     "Controller: QLogic cLOM8214",
     "Interface: PCIe Gen 2 (x8)",
     "Compatibility: HP ProLiant servers",
     "Teaming / VLAN / Jumbo Frame support"],
    img="https://cdn11.bigcommerce.com/s-xdygvn/images/stencil/1280x1280/products/103/169/593717-B21__95200.1667330972.jpg?c=2")

add(35, "IBM 49Y4242 INTEL I340-T4 GIGABIT ETHERNET SERVER ADAPTER", None, 2, 95, MAIN_SERVER, SUB_SRV_NIC, "IBM", COND_ISLENMIS,
    ["Tip: Şəbəkə adapteri (IBM 49Y4242, Intel I340-T4)",
     "Portlar: 4× 10/100/1000 Mbps (RJ-45)",
     "İnterfeys: PCIe (x4)",
     "Uyğunluq: IBM/Lenovo serverlər"],
    ["Тип: Сетевой адаптер (IBM 49Y4242, Intel I340-T4)",
     "Порты: 4× 10/100/1000 Мбит/с (RJ-45)",
     "Интерфейс: PCIe (x4)",
     "Совместимость: серверы IBM/Lenovo"],
    ["Type: Network adapter (IBM 49Y4242, Intel I340-T4)",
     "Ports: 4× 10/100/1000 Mbps (RJ-45)",
     "Interface: PCIe (x4)",
     "Compatibility: IBM/Lenovo servers"],
    img="https://cdn11.bigcommerce.com/s-xdygvn/images/stencil/1280x1280/products/103/169/593717-B21__95200.1667330972.jpg?c=2")

# =================================================================== SFP MODULLAR

add(51, "CISCO SFP-10G-SR 10-2415-03", None, 129, 120, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP+ transceiver (10GBASE-SR)",
     "Dalğa uzunluğu: 850 nm (VCSEL)",
     "Fiber: MMF (LC dupleks)",
     "Məsafə: 300 m (OM3) / 400 m (OM4)",
     "Sürət: 10 Gbps",
     "DOM dəstəyi"],
    ["Тип: SFP+ трансивер (10GBASE-SR)",
     "Длина волны: 850 нм (VCSEL)",
     "Волокно: MMF (LC дуплекс)",
     "Дистанция: 300 м (OM3) / 400 м (OM4)",
     "Скорость: 10 Гбит/с",
     "Поддержка DOM"],
    ["Type: SFP+ transceiver (10GBASE-SR)",
     "Wavelength: 850 nm (VCSEL)",
     "Fiber: MMF (LC duplex)",
     "Reach: 300 m (OM3) / 400 m (OM4)",
     "Data rate: 10 Gbps",
     "DOM support"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(52, "CISCO SFP-10G-SR-S 10-3105-01", None, 14, None, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP+ transceiver (10GBASE-SR, S-Class)",
     "Dalğa uzunluğu: 850 nm",
     "Fiber: MMF (LC dupleks)",
     "Məsafə: 300 m (OM3) / 400 m (OM4)",
     "Sürət: 10 Gbps"],
    ["Тип: SFP+ трансивер (10GBASE-SR, S-Class)",
     "Длина волны: 850 нм",
     "Волокно: MMF (LC дуплекс)",
     "Дистанция: 300 м (OM3) / 400 м (OM4)",
     "Скорость: 10 Гбит/с"],
    ["Type: SFP+ transceiver (10GBASE-SR, S-Class)",
     "Wavelength: 850 nm",
     "Fiber: MMF (LC duplex)",
     "Reach: 300 m (OM3) / 400 m (OM4)",
     "Data rate: 10 Gbps"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(53, "CISCO SFP-GE-T 30-1421-02", None, 43, 30, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP transceiver (1000BASE-T, mis)",
     "Konnetor: RJ-45",
     "Məsafə: 100 m (Cat5e/Cat6)",
     "Sürət: 10/100/1000 Mbps",
     "IEEE 802.3ab"],
    ["Тип: SFP трансивер (1000BASE-T, медь)",
     "Разъём: RJ-45",
     "Дистанция: 100 м (Cat5e/Cat6)",
     "Скорость: 10/100/1000 Мбит/с",
     "IEEE 802.3ab"],
    ["Type: SFP transceiver (1000BASE-T, copper)",
     "Connector: RJ-45",
     "Reach: 100 m (Cat5e/Cat6)",
     "Data rate: 10/100/1000 Mbps",
     "IEEE 802.3ab"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/113308/402446/sfp-ge-t__99049__04506.1787724924.jpg?c=1")

add(54, "AOProvantage SFP-GE-T", None, 41, 25, MAIN_NETWORK, SUB_SFP, "AOProvantage", COND_ISLENMIS,
    ["Tip: SFP transceiver (1000BASE-T, mis)",
     "Konnetor: RJ-45",
     "Məsafə: 100 m",
     "Sürət: 10/100/1000 Mbps"],
    ["Тип: SFP трансивер (1000BASE-T, медь)",
     "Разъём: RJ-45",
     "Дистанция: 100 м",
     "Скорость: 10/100/1000 Мбит/с"],
    ["Type: SFP transceiver (1000BASE-T, copper)",
     "Connector: RJ-45",
     "Reach: 100 m",
     "Data rate: 10/100/1000 Mbps"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/113308/402446/sfp-ge-t__99049__04506.1787724924.jpg?c=1")

add(55, "Finisar SFP FCLF-8521-3", None, 6, 30, MAIN_NETWORK, SUB_SFP, "Finisar", COND_ISLENMIS,
    ["Tip: SFP transceiver (1000BASE-T, mis)",
     "Konnetor: RJ-45",
     "Məsafə: 100 m",
     "Sürət: 10/100/1000 Mbps"],
    ["Тип: SFP трансивер (1000BASE-T, медь)",
     "Разъём: RJ-45",
     "Дистанция: 100 м",
     "Скорость: 10/100/1000 Мбит/с"],
    ["Type: SFP transceiver (1000BASE-T, copper)",
     "Connector: RJ-45",
     "Reach: 100 m",
     "Data rate: 10/100/1000 Mbps"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/113308/402446/sfp-ge-t__99049__04506.1787724924.jpg?c=1")

add(56, "HP 1GB SFP RJ45 MODULE", None, 19, 30, MAIN_NETWORK, SUB_SFP, "HP", COND_ISLENMIS,
    ["Tip: SFP transceiver (1000BASE-T, mis)",
     "Konnetor: RJ-45",
     "Məsafə: 100 m",
     "Sürət: 10/100/1000 Mbps"],
    ["Тип: SFP трансивер (1000BASE-T, медь)",
     "Разъём: RJ-45",
     "Дистанция: 100 м",
     "Скорость: 10/100/1000 Мбит/с"],
    ["Type: SFP transceiver (1000BASE-T, copper)",
     "Connector: RJ-45",
     "Reach: 100 m",
     "Data rate: 10/100/1000 Mbps"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/113308/402446/sfp-ge-t__99049__04506.1787724924.jpg?c=1")

add(57, "CISCO GLC-LH-SMD 1GB", None, 14, 70, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP transceiver (1000BASE-LX/LH)",
     "Dalğa uzunluğu: 1310 nm",
     "Fiber: SMF/MMF (LC dupleks)",
     "Məsafə: 10 km (SMF)",
     "Sürət: 1 Gbps",
     "DOM dəstəyi"],
    ["Тип: SFP трансивер (1000BASE-LX/LH)",
     "Длина волны: 1310 нм",
     "Волокно: SMF/MMF (LC дуплекс)",
     "Дистанция: 10 км (SMF)",
     "Скорость: 1 Гбит/с",
     "Поддержка DOM"],
    ["Type: SFP transceiver (1000BASE-LX/LH)",
     "Wavelength: 1310 nm",
     "Fiber: SMF/MMF (LC duplex)",
     "Reach: 10 km (SMF)",
     "Data rate: 1 Gbps",
     "DOM support"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/49155/572438/cisco-glc-lh-smd_1-1__59978.1777570167.jpg?c=1")

add(58, "CISCO SFP-10G-LR SM", None, 14, 120, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP+ transceiver (10GBASE-LR)",
     "Dalğa uzunluğu: 1310 nm",
     "Fiber: SMF (LC dupleks)",
     "Məsafə: 10 km",
     "Sürət: 10 Gbps"],
    ["Тип: SFP+ трансивер (10GBASE-LR)",
     "Длина волны: 1310 нм",
     "Волокно: SMF (LC дуплекс)",
     "Дистанция: 10 км",
     "Скорость: 10 Гбит/с"],
    ["Type: SFP+ transceiver (10GBASE-LR)",
     "Wavelength: 1310 nm",
     "Fiber: SMF (LC duplex)",
     "Reach: 10 km",
     "Data rate: 10 Gbps"],
    img="https://img.genuinemodules.com/cache/catalog/products/SFP-10G-LR/SFP-10G-LR-1-800x800.jpg")

add(59, "HP 8G SW FC SFP+ AJ817A", None, 4, 65, MAIN_NETWORK, SUB_SFP, "HP", COND_ISLENMIS,
    ["Tip: SFP+ transceiver (8G Fibre Channel SW)",
     "Sürət: 8 Gbps FC",
     "Fiber: MMF (LC)",
     "Uyğunluq: HP SAN kommutatorları"],
    ["Тип: SFP+ трансивер (8G Fibre Channel SW)",
     "Скорость: 8 Гбит/с FC",
     "Волокно: MMF (LC)",
     "Совместимость: SAN-коммутаторы HP"],
    ["Type: SFP+ transceiver (8G Fibre Channel SW)",
     "Data rate: 8 Gbps FC",
     "Fiber: MMF (LC)",
     "Compatibility: HP SAN switches"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(60, "CISCO TWINGIG CVR-X2-SFP V02", None, 6, 50, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: TwinGig konvertor modulu (CVR-X2-SFP)",
     "X2 → 2× SFP çevirici",
     "Uyğunluq: Cisco Catalyst X2 slotları"],
    ["Тип: Конвертерный модуль TwinGig (CVR-X2-SFP)",
     "Преобразование X2 → 2× SFP",
     "Совместимость: слоты Cisco Catalyst X2"],
    ["Type: TwinGig converter module (CVR-X2-SFP)",
     "X2 → 2× SFP conversion",
     "Compatibility: Cisco Catalyst X2 slots"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(61, "CISCO DS-SFP-FC 16G-SW", None, 20, 55, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP+ transceiver (16G Fibre Channel SW)",
     "Sürət: 16 Gbps FC",
     "Fiber: MMF (LC)",
     "Uyğunluq: Cisco MDS kommutatorları"],
    ["Тип: SFP+ трансивер (16G Fibre Channel SW)",
     "Скорость: 16 Гбит/с FC",
     "Волокно: MMF (LC)",
     "Совместимость: коммутаторы Cisco MDS"],
    ["Type: SFP+ transceiver (16G Fibre Channel SW)",
     "Data rate: 16 Gbps FC",
     "Fiber: MMF (LC)",
     "Compatibility: Cisco MDS switches"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(62, "CISCO XFP-10G LR 10-1989-03", None, 4, 75, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: XFP transceiver (10GBASE-LR)",
     "Dalğa uzunluğu: 1310 nm",
     "Fiber: SMF (LC)",
     "Məsafə: 10 km",
     "Sürət: 10 Gbps"],
    ["Тип: XFP трансивер (10GBASE-LR)",
     "Длина волны: 1310 нм",
     "Волокно: SMF (LC)",
     "Дистанция: 10 км",
     "Скорость: 10 Гбит/с"],
    ["Type: XFP transceiver (10GBASE-LR)",
     "Wavelength: 1310 nm",
     "Fiber: SMF (LC)",
     "Reach: 10 km",
     "Data rate: 10 Gbps"],
    img="https://cdn11.bigcommerce.com/s-e692hdujm7/images/stencil/500x500/products/4240/4458/SFP10GLR-Ra__77044.1701880950.jpg?c=2")

add(63, "CISCO GLC-FE-100 SFP", None, 100, 20, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP transceiver (100BASE-FX)",
     "Dalğa uzunluğu: 1310 nm",
     "Fiber: MMF (LC)",
     "Məsafə: 2 km",
     "Sürət: 100 Mbps"],
    ["Тип: SFP трансивер (100BASE-FX)",
     "Длина волны: 1310 нм",
     "Волокно: MMF (LC)",
     "Дистанция: 2 км",
     "Скорость: 100 Мбит/с"],
    ["Type: SFP transceiver (100BASE-FX)",
     "Wavelength: 1310 nm",
     "Fiber: MMF (LC)",
     "Reach: 2 km",
     "Data rate: 100 Mbps"],
    img="https://cdn11.bigcommerce.com/s-gtdjhqa/images/stencil/160w/products/31/3740/GMFP083-VSL05D_ThreeQuarter__09692.1617725245.jpg?c=2")

add(64, "CISCO DS-SFP-FC4G SW", None, 16, 40, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP transceiver (4G Fibre Channel SW)",
     "Sürət: 4 Gbps FC",
     "Fiber: MMF (LC)",
     "Uyğunluq: Cisco MDS kommutatorları"],
    ["Тип: SFP трансивер (4G Fibre Channel SW)",
     "Скорость: 4 Гбит/с FC",
     "Волокно: MMF (LC)",
     "Совместимость: коммутаторы Cisco MDS"],
    ["Type: SFP transceiver (4G Fibre Channel SW)",
     "Data rate: 4 Gbps FC",
     "Fiber: MMF (LC)",
     "Compatibility: Cisco MDS switches"],
    img="https://cdn11.bigcommerce.com/s-gtdjhqa/images/stencil/160w/products/31/3740/GMFP083-VSL05D_ThreeQuarter__09692.1617725245.jpg?c=2")

add(65, "CISCO GLC-SX-MM 1G", None, 245, 30, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP transceiver (1000BASE-SX)",
     "Dalğa uzunluğu: 850 nm (VCSEL)",
     "Fiber: MMF (LC dupleks)",
     "Məsafə: 550 m (OM2)",
     "Sürət: 1 Gbps"],
    ["Тип: SFP трансивер (1000BASE-SX)",
     "Длина волны: 850 нм (VCSEL)",
     "Волокно: MMF (LC дуплекс)",
     "Дистанция: 550 м (OM2)",
     "Скорость: 1 Гбит/с"],
    ["Type: SFP transceiver (1000BASE-SX)",
     "Wavelength: 850 nm (VCSEL)",
     "Fiber: MMF (LC duplex)",
     "Reach: 550 m (OM2)",
     "Data rate: 1 Gbps"],
    img="https://cdn11.bigcommerce.com/s-gtdjhqa/images/stencil/160w/products/31/3740/GMFP083-VSL05D_ThreeQuarter__09692.1617725245.jpg?c=2")

add(66, "CISCO FET-10G SFP", None, 4, 110, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP+ transceiver (Fabric Extender 10G)",
     "Sürət: 10 Gbps",
     "Fiber: MMF",
     "Uyğunluq: Cisco Nexus Fabric Extender"],
    ["Тип: SFP+ трансивер (Fabric Extender 10G)",
     "Скорость: 10 Гбит/с",
     "Волокно: MMF",
     "Совместимость: Cisco Nexus Fabric Extender"],
    ["Type: SFP+ transceiver (Fabric Extender 10G)",
     "Data rate: 10 Gbps",
     "Fiber: MMF",
     "Compatibility: Cisco Nexus Fabric Extender"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(68, "CISCO X2-10GB-SR SFP", None, 11, 70, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: X2 transceiver (10GBASE-SR)",
     "Dalğa uzunluğu: 850 nm",
     "Fiber: MMF (SC)",
     "Məsafə: 300 m",
     "Sürət: 10 Gbps"],
    ["Тип: X2 трансивер (10GBASE-SR)",
     "Длина волны: 850 нм",
     "Волокно: MMF (SC)",
     "Дистанция: 300 м",
     "Скорость: 10 Гбит/с"],
    ["Type: X2 transceiver (10GBASE-SR)",
     "Wavelength: 850 nm",
     "Fiber: MMF (SC)",
     "Reach: 300 m",
     "Data rate: 10 Gbps"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(69, "CISCO DS-SFP-FC8G-SW", None, 41, 50, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: SFP+ transceiver (8G Fibre Channel SW)",
     "Sürət: 8 Gbps FC",
     "Fiber: MMF (LC)",
     "Uyğunluq: Cisco MDS kommutatorları"],
    ["Тип: SFP+ трансивер (8G Fibre Channel SW)",
     "Скорость: 8 Гбит/с FC",
     "Волокно: MMF (LC)",
     "Совместимость: коммутаторы Cisco MDS"],
    ["Type: SFP+ transceiver (8G Fibre Channel SW)",
     "Data rate: 8 Gbps FC",
     "Fiber: MMF (LC)",
     "Compatibility: Cisco MDS switches"],
    img="https://cdn11.bigcommerce.com/s-b9xom8/images/stencil/1280x1280/products/221838/326145/1024204086__27843.1782380622.jpg?c=2")

add(70, "CISCO  XENPAK 10GB-LR SFP", None, 3, 120, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: XENPAK transceiver (10GBASE-LR)",
     "Dalğa uzunluğu: 1310 nm",
     "Fiber: SMF",
     "Məsafə: 10 km",
     "Sürət: 10 Gbps"],
    ["Тип: XENPAK трансивер (10GBASE-LR)",
     "Длина волны: 1310 нм",
     "Волокно: SMF",
     "Дистанция: 10 км",
     "Скорость: 10 Гбит/с"],
    ["Type: XENPAK transceiver (10GBASE-LR)",
     "Wavelength: 1310 nm",
     "Fiber: SMF",
     "Reach: 10 km",
     "Data rate: 10 Gbps"],
    img="https://cdn11.bigcommerce.com/s-e692hdujm7/images/stencil/500x500/products/4240/4458/SFP10GLR-Ra__77044.1701880950.jpg?c=2")

add(71, "CISCO QSFP 40GB-SR4-S", None, 20, 200, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: QSFP+ transceiver (40GBASE-SR4)",
     "Dalğa uzunluğu: 850 nm",
     "Fiber: MMF (MPO-12)",
     "Məsafə: 100 m (OM3) / 150 m (OM4)",
     "Sürət: 40 Gbps"],
    ["Тип: QSFP+ трансивер (40GBASE-SR4)",
     "Длина волны: 850 нм",
     "Волокно: MMF (MPO-12)",
     "Дистанция: 100 м (OM3) / 150 м (OM4)",
     "Скорость: 40 Гбит/с"],
    ["Type: QSFP+ transceiver (40GBASE-SR4)",
     "Wavelength: 850 nm",
     "Fiber: MMF (MPO-12)",
     "Reach: 100 m (OM3) / 150 m (OM4)",
     "Data rate: 40 Gbps"],
    img="https://img.genuinemodules.com/cache/catalog/products/QSFP-40G-SR4-S/QSFP-40G-SR4-S-1-800x800.jpg")

add(72, "CISCO QSFP 100GB-SR4-S", None, 7, 220, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: QSFP28 transceiver (100GBASE-SR4)",
     "Dalğa uzunluğu: 850 nm",
     "Fiber: MMF (MPO-12)",
     "Məsafə: 70 m (OM3) / 100 m (OM4)",
     "Sürət: 100 Gbps"],
    ["Тип: QSFP28 трансивер (100GBASE-SR4)",
     "Длина волны: 850 нм",
     "Волокно: MMF (MPO-12)",
     "Дистанция: 70 м (OM3) / 100 м (OM4)",
     "Скорость: 100 Гбит/с"],
    ["Type: QSFP28 transceiver (100GBASE-SR4)",
     "Wavelength: 850 nm",
     "Fiber: MMF (MPO-12)",
     "Reach: 70 m (OM3) / 100 m (OM4)",
     "Data rate: 100 Gbps"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/401564/974043/Cisco-QSFP-100G-SR4-S__98259.1784580373.jpg?c=1")

add(73, "CISCO QSFP 100GB-SM-SR", None, 9, 220, MAIN_NETWORK, SUB_SFP, "Cisco", COND_ISLENMIS,
    ["Tip: QSFP28 transceiver (100GBASE-SR)",
     "Dalğa uzunluğu: 850 nm",
     "Fiber: MMF",
     "Məsafə: 100 m (OM4)",
     "Sürət: 100 Gbps"],
    ["Тип: QSFP28 трансивер (100GBASE-SR)",
     "Длина волны: 850 нм",
     "Волокно: MMF",
     "Дистанция: 100 м (OM4)",
     "Скорость: 100 Гбит/с"],
    ["Type: QSFP28 transceiver (100GBASE-SR)",
     "Wavelength: 850 nm",
     "Fiber: MMF",
     "Reach: 100 m (OM4)",
     "Data rate: 100 Gbps"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/401564/974043/Cisco-QSFP-100G-SR4-S__98259.1784580373.jpg?c=1")

# ================================================================= SERVER AKSESUARLARI

add(9, "DELL HDD 2.5 CADDY", None, 5, 30, MAIN_SERVER, SUB_SRV_ACC, "Dell", COND_ISLENMIS,
    ["Tip: Sürücü qəfəsi (Dell, 2.5\" SFF)",
     "İnterfeys: SAS/SATA",
     "Uyğunluq: Dell PowerEdge serverlər (12-14-cü nəsil)"],
    ["Тип: Корзина для диска (Dell, 2.5\" SFF)",
     "Интерфейс: SAS/SATA",
     "Совместимость: серверы Dell PowerEdge (12-14 поколение)"],
    ["Type: Drive caddy (Dell, 2.5\" SFF)",
     "Interface: SAS/SATA",
     "Compatibility: Dell PowerEdge servers (12th-14th gen)"],
    img="https://cdn11.bigcommerce.com/s-qfzamxn9kz/images/stencil/original/products/161984/222239/dell-kf248-caddy-tray-for-hard-disk-drive-hdd-5__39187.1534979524.jpg?c=2")

add(10, "HP SMART STORAGE BATTERY 815983-001", None, 1, 65, MAIN_SERVER, SUB_SRV_ACC, "HP", COND_ISLENMIS,
    ["Tip: RAID kontroller batareyası (HP Smart Array)",
     "Part nömrəsi: 815983-001",
     "Uyğunluq: HP P440/P840 RAID kontrollerlər"],
    ["Тип: Батарея RAID-контроллера (HP Smart Array)",
     "Номер детали: 815983-001",
     "Совместимость: RAID-контроллеры HP P440/P840"],
    ["Type: RAID controller battery (HP Smart Array)",
     "Part number: 815983-001",
     "Compatibility: HP P440/P840 RAID controllers"],
    img="https://cdn11.bigcommerce.com/s-xdygvn/images/stencil/1280x1280/products/2689/4392/HPE_727258-B21__00334.1615944902.jpg?c=2")

add(18, "HP 3.5 SAS CADDY", None, 4, 30, MAIN_SERVER, SUB_SRV_ACC, "HP", COND_ISLENMIS,
    ["Tip: Sürücü qəfəsi (HP, 3.5\" LFF)",
     "İnterfeys: SAS/SATA",
     "Uyğunluq: HP ProLiant serverlər"],
    ["Тип: Корзина для диска (HP, 3.5\" LFF)",
     "Интерфейс: SAS/SATA",
     "Совместимость: серверы HP ProLiant"],
    ["Type: Drive caddy (HP, 3.5\" LFF)",
     "Interface: SAS/SATA",
     "Compatibility: HP ProLiant servers"],
    img="https://cdn11.bigcommerce.com/s-qfzamxn9kz/images/stencil/original/products/161984/222239/dell-kf248-caddy-tray-for-hard-disk-drive-hdd-5__39187.1534979524.jpg?c=2")

add(29, "DELL BATTERY MODULE T40JJ 3.6V", None, 5, 35, MAIN_SERVER, SUB_SRV_ACC, "Dell", COND_ISLENMIS,
    ["Tip: RAID kontroller batareyası (Dell)",
     "Gərginlik: 3.6 V",
     "Part nömrəsi: T40JJ"],
    ["Тип: Батарея RAID-контроллера (Dell)",
     "Напряжение: 3.6 В",
     "Номер детали: T40JJ"],
    ["Type: RAID controller battery (Dell)",
     "Voltage: 3.6 V",
     "Part number: T40JJ"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/105619/664059/T40JJ__01300.1784579676.jpg?c=1")

add(49, "HP SERVER FAN PFM0412HS", None, 4, 30, MAIN_SERVER, SUB_SRV_ACC, "HP", COND_ISLENMIS,
    ["Tip: Server ventilyatoru (HP PFM0412HS)",
     "Ölçü: 40 mm",
     "Uyğunluq: HP ProLiant serverlər"],
    ["Тип: Серверный вентилятор (HP PFM0412HS)",
     "Размер: 40 мм",
     "Совместимость: серверы HP ProLiant"],
    ["Type: Server fan (HP PFM0412HS)",
     "Size: 40 mm",
     "Compatibility: HP ProLiant servers"],
    img="https://cdn11.bigcommerce.com/s-qfzamxn9kz/images/stencil/original/products/143398/375980/750688-001__11506.1562085396.jpg?c=2")

add(50, "HP SERVER FAN GFM0412SS", None, 2, 30, MAIN_SERVER, SUB_SRV_ACC, "HP", COND_ISLENMIS,
    ["Tip: Server ventilyatoru (HP GFM0412SS)",
     "Ölçü: 40 mm",
     "Uyğunluq: HP ProLiant serverlər"],
    ["Тип: Серверный вентилятор (HP GFM0412SS)",
     "Размер: 40 мм",
     "Совместимость: серверы HP ProLiant"],
    ["Type: Server fan (HP GFM0412SS)",
     "Size: 40 mm",
     "Compatibility: HP ProLiant servers"],
    img="https://cdn11.bigcommerce.com/s-b18w/images/stencil/1280x1280/products/39015/96233/667882-001_Front__91291.1679416468.1280.1280__05121.1679428830.jpg?c=2")

# ========================================================================= HDD

add(12, "WESTERN DIGITAL WD8004FRYZ SATA 6GB/S DC HA750 P/N 2W10441", None, 1, None, MAIN_SERVER, SUB_HDD, "Western Digital", COND_ISLENMIS,
    ["Tip: Server HDD (Western Digital Gold)",
     "Tutum: 8 TB",
     "İnterfeys: SATA 6 Gb/s",
     "Sürət: 7200 RPM",
     "Keş: 256 MB",
     "Format: 3.5\" (512e)",
     "MTBF: 2.5 milyon saat"],
    ["Тип: Серверный HDD (Western Digital Gold)",
     "Ёмкость: 8 ТБ",
     "Интерфейс: SATA 6 Гбит/с",
     "Скорость: 7200 об/мин",
     "Кэш: 256 МБ",
     "Формат: 3.5\" (512e)",
     "MTBF: 2.5 млн часов"],
    ["Type: Server HDD (Western Digital Gold)",
     "Capacity: 8 TB",
     "Interface: SATA 6 Gb/s",
     "Speed: 7200 RPM",
     "Cache: 256 MB",
     "Form factor: 3.5\" (512e)",
     "MTBF: 2.5M hours"],
    img="https://esaitech.com/cdn/shop/files/5_2cb6799b-e8d5-4842-92fc-b40c2c50abde_1024x.jpg?v=1771982469")

# ========================================================== ENERJİ TƏCHİZATI (server + şəbəkə)

add(40, "CISCO PWR-C1-1100WAG", None, 16, 190, MAIN_NETWORK, SUB_NET_PSU, "Cisco", COND_ISLENMIS,
    ["Tip: Switch enerji təchizatı (Cisco PWR-C1-1100WAC)",
     "Güc: 1100 W",
     "Giriş: 115-240 VAC (autoranging)",
     "Uyğunluq: Cisco Catalyst 3850 seriyası",
     "Hot-swappable"],
    ["Тип: Блок питания коммутатора (Cisco PWR-C1-1100WAC)",
     "Мощность: 1100 Вт",
     "Вход: 115-240 VAC (автодиапазон)",
     "Совместимость: серия Cisco Catalyst 3850",
     "Горячая замена"],
    ["Type: Switch power supply (Cisco PWR-C1-1100WAC)",
     "Power: 1100 W",
     "Input: 115-240 VAC (autoranging)",
     "Compatibility: Cisco Catalyst 3850 series",
     "Hot-swappable"],
    img="https://www.networktigers.com/cdn/shop/files/cisco-PWR-C1-1100WAC_463x.progressive.jpg?v=1685985937")

add(41, "CISCO C3KX-PWR-715WAG", None, 5, 150, MAIN_NETWORK, SUB_NET_PSU, "Cisco", COND_ISLENMIS,
    ["Tip: Switch enerji təchizatı (Cisco C3KX-PWR-715WAC)",
     "Güc: 715 W",
     "Uyğunluq: Cisco Catalyst 3560-X / 3750-X",
     "Hot-swappable"],
    ["Тип: Блок питания коммутатора (Cisco C3KX-PWR-715WAC)",
     "Мощность: 715 Вт",
     "Совместимость: Cisco Catalyst 3560-X / 3750-X",
     "Горячая замена"],
    ["Type: Switch power supply (Cisco C3KX-PWR-715WAC)",
     "Power: 715 W",
     "Compatibility: Cisco Catalyst 3560-X / 3750-X",
     "Hot-swappable"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/104701/401323/C3KX-PWR-715WAC__15225__47784.1777568822.jpg?c=1")

add(42, "CISCO C3KX-PWR-350WAG", None, 10, 120, MAIN_NETWORK, SUB_NET_PSU, "Cisco", COND_ISLENMIS,
    ["Tip: Switch enerji təchizatı (Cisco C3KX-PWR-350WAC)",
     "Güc: 350 W",
     "Uyğunluq: Cisco Catalyst 3560-X / 3750-X",
     "Hot-swappable"],
    ["Тип: Блок питания коммутатора (Cisco C3KX-PWR-350WAC)",
     "Мощность: 350 Вт",
     "Совместимость: Cisco Catalyst 3560-X / 3750-X",
     "Горячая замена"],
    ["Type: Switch power supply (Cisco C3KX-PWR-350WAC)",
     "Power: 350 W",
     "Compatibility: Cisco Catalyst 3560-X / 3750-X",
     "Hot-swappable"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/104701/401323/C3KX-PWR-715WAC__15225__47784.1777568822.jpg?c=1")

add(43, "HP DPS-500AB-13A POWER SUPPLY", None, 5, 40, MAIN_SERVER, SUB_SRV_PSU, "HP", COND_ISLENMIS,
    ["Tip: Server enerji təchizatı (HP DPS-500AB-13A)",
     "Güc: 500 W",
     "Effektivlik: 80 Plus Platinum",
     "Uyğunluq: HP DL360/DL380/ML350 Gen9",
     "Hot-plug"],
    ["Тип: Блок питания сервера (HP DPS-500AB-13A)",
     "Мощность: 500 Вт",
     "КПД: 80 Plus Platinum",
     "Совместимость: HP DL360/DL380/ML350 Gen9",
     "Горячая замена"],
    ["Type: Server power supply (HP DPS-500AB-13A)",
     "Power: 500 W",
     "Efficiency: 80 Plus Platinum",
     "Compatibility: HP DL360/DL380/ML350 Gen9",
     "Hot-plug"],
    img="https://cdn11.bigcommerce.com/s-wb5rkphzy3/images/stencil/1280x1280/products/29582/433193/dps-500ab-13a__20430__48089.1777570123.jpg?c=1")

add(44, "HP DPS-460MB-A POWER SUPPLY", None, 5, 35, MAIN_SERVER, SUB_SRV_PSU, "HP", COND_ISLENMIS,
    ["Tip: Server enerji təchizatı (HP DPS-460MB)",
     "Güc: 460 W",
     "Effektivlik: 94% (Platinum Plus)",
     "Uyğunluq: HP ProLiant Gen8 serverlər",
     "Hot-plug"],
    ["Тип: Блок питания сервера (HP DPS-460MB)",
     "Мощность: 460 Вт",
     "КПД: 94% (Platinum Plus)",
     "Совместимость: серверы HP ProLiant Gen8",
     "Горячая замена"],
    ["Type: Server power supply (HP DPS-460MB)",
     "Power: 460 W",
     "Efficiency: 94% (Platinum Plus)",
     "Compatibility: HP ProLiant Gen8 servers",
     "Hot-plug"],
    img="https://cdn11.bigcommerce.com/s-g6oxherh18/images/stencil/original/products/61382/49114/main-squared__33454.1780371324.jpg?c=2")

add(45, "HP PS-3381-1C1 400W", None, 1, 35, MAIN_SERVER, SUB_SRV_PSU, "HP", COND_ISLENMIS,
    ["Tip: Server enerji təchizatı (HP PS-3381-1C1)",
     "Güc: 400 W",
     "Hot-swap"],
    ["Тип: Блок питания сервера (HP PS-3381-1C1)",
     "Мощность: 400 Вт",
     "Горячая замена"],
    ["Type: Server power supply (HP PS-3381-1C1)",
     "Power: 400 W",
     "Hot-swap"],
    img="https://cdn11.bigcommerce.com/s-g6oxherh18/images/stencil/original/products/61382/49114/main-squared__33454.1780371324.jpg?c=2")

add(46, "HP ESP128 325W POWER SUPPLY", None, 1, 35, MAIN_SERVER, SUB_SRV_PSU, "HP", COND_ISLENMIS,
    ["Tip: Server enerji təchizatı (HP ESP128)",
     "Güc: 325 W"],
    ["Тип: Блок питания сервера (HP ESP128)",
     "Мощность: 325 Вт"],
    ["Type: Server power supply (HP ESP128)",
     "Power: 325 W"],
    img="https://cdn11.bigcommerce.com/s-g6oxherh18/images/stencil/original/products/61382/49114/main-squared__33454.1780371324.jpg?c=2")

# ==================================================================== ŞƏBƏKƏ VENTİLYATORLARI (aksesuarlar)

add(47, "CISCO C3KX-FAN-23CFM", None, 16, 30, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Switch ventilyator modulu (Cisco C3KX-FAN-23CFM)",
     "Hava axını: 23 CFM",
     "Uyğunluq: Cisco Catalyst 3560-X / 3750-X",
     "Hot-swappable"],
    ["Тип: Модуль вентилятора коммутатора (Cisco C3KX-FAN-23CFM)",
     "Воздушный поток: 23 CFM",
     "Совместимость: Cisco Catalyst 3560-X / 3750-X",
     "Горячая замена"],
    ["Type: Switch fan module (Cisco C3KX-FAN-23CFM)",
     "Airflow: 23 CFM",
     "Compatibility: Cisco Catalyst 3560-X / 3750-X",
     "Hot-swappable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/399334/765346/C3KX-FAN-23CFM-lg__79532.1784325866.jpg?c=1")

add(48, "CISCO C3850 FAN-T1", None, 15, 35, MAIN_NETWORK, SUB_NET_ACC, "Cisco", COND_ISLENMIS,
    ["Tip: Switch ventilyator modulu (Cisco C3850 FAN-T1)",
     "Uyğunluq: Cisco Catalyst 3850 seriyası",
     "Hot-swappable"],
    ["Тип: Модуль вентилятора коммутатора (Cisco C3850 FAN-T1)",
     "Совместимость: серия Cisco Catalyst 3850",
     "Горячая замена"],
    ["Type: Switch fan module (Cisco C3850 FAN-T1)",
     "Compatibility: Cisco Catalyst 3850 series",
     "Hot-swappable"],
    img="https://cdn11.bigcommerce.com/s-4jpol1blth/images/stencil/600x600/products/399334/765346/C3KX-FAN-23CFM-lg__79532.1784325866.jpg?c=1")

# =================================================================== UPS AKSESUARLARI

add(27, "EATON NMC SBP-00154 UPS NETWORK CARD", None, 1, 60, MAIN_UPS, SUB_UPS_ACC, "Eaton", COND_ISLENMIS,
    ["Tip: UPS şəbəkə kartı (Eaton Network-MS / SBP-00154)",
     "Slot: Mini-Slot",
     "Protokol: SNMP v1/v3, HTTP/HTTPS, IPv6",
     "Ethernet: 10/100BaseT",
     "MIB: RFC 1628, Eaton Pulsar MIB"],
    ["Тип: Сетевая карта ИБП (Eaton Network-MS / SBP-00154)",
     "Слот: Mini-Slot",
     "Протоколы: SNMP v1/v3, HTTP/HTTPS, IPv6",
     "Ethernet: 10/100BaseT",
     "MIB: RFC 1628, Eaton Pulsar MIB"],
    ["Type: UPS network card (Eaton Network-MS / SBP-00154)",
     "Slot: Mini-Slot",
     "Protocols: SNMP v1/v3, HTTP/HTTPS, IPv6",
     "Ethernet: 10/100BaseT",
     "MIB: RFC 1628, Eaton Pulsar MIB"],
    img="https://media.stockinthechannel.com/pic/86XMzZfPgka5wWQrinqC3A.r.jpg")

add(28, "MGE UPS CARD YUPI - 34003630 00 IB-NT:14", None, 1, 45, MAIN_UPS, SUB_UPS_ACC, "MGE", COND_ISLENMIS,
    ["Tip: UPS şəbəkə kartı (MGE NMC)",
     "Part nömrəsi: 34003630 00",
     "Slot: Mini-Slot",
     "Protokol: SNMP, Web monitorinq"],
    ["Тип: Сетевая карта ИБП (MGE NMC)",
     "Номер детали: 34003630 00",
     "Слот: Mini-Slot",
     "Протоколы: SNMP, Web-мониторинг"],
    ["Type: UPS network card (MGE NMC)",
     "Part number: 34003630 00",
     "Slot: Mini-Slot",
     "Protocols: SNMP, Web monitoring"],
    img="https://secure.ups-trader.co.uk/3388-thickbox_default/eaton-mge-66244-nic-card.jpg")

# ------------------------------------------------------------ workbook ----
HEADERS = [
    ("№", 6),
    ("Model", 34),
    ("Barkod", 16),
    ("Vəziyyəti", 12),
    ("Qty", 6),
    ("Qiymət (AZN) — endirimli", 18),
    ("Köhnə qiymət (AZN)", 18),
    ("Əsas kateqoriya", 24),
    ("Alt kateqoriya", 22),
    ("Brend", 16),
    ("Xüsusiyyətlər (AZ)", 44),
    ("Xüsusiyyətlər (RU)", 44),
    ("Xüsusiyyətlər (EN)", 44),
    ("Şəkil linkləri", 60),
]

CATEGORY_SHEET = [
    (MAIN_NETWORK, "sebeke-avadanliqlari", "Kommutator", "kommutator", "Mövcud"),
    (MAIN_NETWORK, "sebeke-avadanliqlari", "Router", "router", "Mövcud"),
    (MAIN_NETWORK, "sebeke-avadanliqlari", "VoIP və IP telefonlar", "voip-ip-telefonlar", "Mövcud"),
    (MAIN_NETWORK, "sebeke-avadanliqlari", "Şəbəkə aksesuarları", "sebeke-aksesuarlari", "Mövcud"),
    (MAIN_NETWORK, "sebeke-avadanliqlari", "SFP modullar", "sfp-modullar", "Mövcud"),
    (MAIN_NETWORK, "sebeke-avadanliqlari", "Şəbəkə enerji təchizatı", "sebeke-enerji-techizati", "YENİ — yaradılmalıdır"),
    (MAIN_SERVER, "server", "Şəbəkə adapteri", "server-sebeke-adapteri", "Mövcud"),
    (MAIN_SERVER, "server", "HDD", "server-hdd", "Mövcud"),
    (MAIN_SERVER, "server", "Server aksesuarları", "server-aksesuarlari", "Mövcud"),
    (MAIN_SERVER, "server", "Server enerji təchizatı", "server-enerji-techizati", "YENİ — yaradılmalıdır"),
    (MAIN_UPS, "ups", "UPS aksesuarları", "ups-aksesuarlari", "Mövcud"),
]


def build_wb():
    wb = Workbook()

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # ---------- Sheet 1: Məhsullar ----------
    ws = wb.active
    ws.title = "Məhsullar"

    ws["A1"] = "Diger (HP / Aruba / Cisco / NETGEAR / H3C / Dell / WD / IBM / Eaton) — Məhsul siyahısı (31.08.2026) | IT Market üçün"
    ws["A1"].font = title_font
    ws["A2"] = (
        "Qiymət (AZN) hazırkı endirimli qiymətdir və dəyişmir. Köhnə qiymət = Qiymət + 55 AZN. "
        "Vəziyyəti: \"Yeni\" yazılan məhsullar yeni, digərləri işlənmişdir. "
        "RU və EN xüsusiyyətlər lüğət bazasına əlavə olunmaq üçündür."
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")

    header_row = 3
    for col, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    row = header_row + 1
    for num, model, barcode, qty, price, main, sub, brand, cond, az, ru, en in P:
        old_price = round(price + 55, 2) if price is not None else None
        values = [
            num, model, barcode or "", cond, qty,
            price if price is not None else "",
            old_price if old_price is not None else "",
            main, sub, brand,
            "\n".join(az), "\n".join(ru), "\n".join(en),
            IMAGES.get(num, ""),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border
            cell.alignment = center if col in (1, 4, 5, 6, 7) else wrap
        row += 1

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{row - 1}"

    # ---------- Sheet 2: Kateqoriyalar ----------
    ws2 = wb.create_sheet("Kateqoriyalar")
    ws2["A1"] = "Saytda mövcud və yeni təklif olunan kateqoriyalar"
    ws2["A1"].font = title_font

    cat_headers = ["Əsas kateqoriya", "Əsas slug", "Alt kateqoriya", "Alt slug", "Status"]
    for col, name in enumerate(cat_headers, start=1):
        cell = ws2.cell(row=3, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = 34

    for r, (main, main_slug, sub, sub_slug, status) in enumerate(CATEGORY_SHEET, start=4):
        for c, v in enumerate([main, main_slug, sub, sub_slug, status], start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = wrap
            if status.startswith("YENİ"):
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
                cell.font = Font(bold=True, color="7F6000")

    ws2.freeze_panes = "A4"

    out = "Sayt üçün 31082026- DIGER_260831_110926.xlsx"
    wb.save(out)
    print(f"OK: {out} ({len(P)} products)")


if __name__ == "__main__":
    build_wb()
